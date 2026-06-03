"""
Excel → Markdown vault 轉換器（Karpathy LLM Wiki 風格）

讀取 各單位FAQ/FAQ更新作業(*)_iASK版_*.xlsx
產出完整 karpathy-wiki 結構：
  vault/
    index.md                  ← 根目錄總覽
    log.md                    ← 變動編年史（append-only）
    <DEPT>/_index.md          ← 部門目錄
    <DEPT>/[ID]-slug.md       ← 一檔一 Q，含 frontmatter + Related 跨頁引用
"""
from __future__ import annotations

import argparse
import datetime as _dt
import os
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

import openpyxl


HERE = Path(__file__).resolve().parent
ROOT = HERE.parent  # poc/
SOURCE_DIR = ROOT.parent / "各單位FAQ"
VAULT_DIR = ROOT / "vault"


# 每個檔的「正確 sheet」對應（首選；若不存在則 fallback 到第一張）
SHEET_OVERRIDES = {
    "FAQ更新作業(PMC)_iASK版_20251117.xlsx": "FAQ (1023)",
}

ID_PATTERN = re.compile(r"\[\s*([A-Z]{2,4}\d{2,4})\s*\]")


@dataclass
class FAQItem:
    dept: str
    id: str
    question: str
    answer: str
    links: list[dict] = field(default_factory=list)
    source_file: str = ""
    sheet: str = ""
    row: int = 0


def detect_dept(filename: str) -> str:
    m = re.search(r"FAQ更新作業\(([A-Z]+)\)", filename)
    return m.group(1) if m else "UNKNOWN"


def pick_sheet(wb: openpyxl.Workbook, filename: str) -> str:
    override = SHEET_OVERRIDES.get(filename)
    if override and override in wb.sheetnames:
        return override
    # 首選叫「FAQ上稿內容」（含變體）的 sheet
    for sn in wb.sheetnames:
        if "FAQ上稿內容" in sn:
            return sn
    return wb.sheetnames[0]


def parse_links(raw: str | None) -> list[dict]:
    """
    連結欄常見格式：
      '名稱;url'
      '名稱:https://...'
      '名稱;url\\n名稱2;url2'
    回傳 [{name, url}, ...]
    """
    if not raw:
        return []
    text = str(raw).strip()
    if not text:
        return []
    items: list[dict] = []
    for line in text.split("\n"):
        line = line.strip()
        if not line:
            continue
        # 找 URL
        url_match = re.search(r"https?://\S+", line)
        if not url_match:
            # 沒 URL 就只記名稱
            items.append({"name": line, "url": ""})
            continue
        url = url_match.group(0).rstrip("，,。)」】")
        name = line[: url_match.start()].strip().rstrip(":：;；").strip()
        if not name:
            name = url
        items.append({"name": name, "url": url})
    return items


def extract_id_and_title(raw_q: str) -> tuple[str, str]:
    """
    Q 欄位常見：
      '[PMC005] 何時提交Pipeline？如何提交 ?'
      '[CFC001] xxx\\n原公告日期: 2024/4/25'
    """
    text = str(raw_q or "").strip()
    if not text:
        return "", ""
    # 拆掉「原公告日期」之類的雜訊
    first_line = text.split("\n", 1)[0].strip()
    m = ID_PATTERN.search(first_line)
    if not m:
        return "", first_line
    qid = m.group(1)
    title = first_line[m.end() :].strip().lstrip(":：")
    return qid, title


def slugify(s: str, maxlen: int = 30) -> str:
    s = re.sub(r"[\s　]+", "-", s.strip())
    s = re.sub(r"[\\/:*?\"<>|]", "", s)
    s = s.strip("-")
    return s[:maxlen]


def iter_faq_rows(ws) -> Iterable[tuple[int, tuple]]:
    """逐列回傳 (row_index, (Q, A, links))，只取前 3 欄。"""
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if idx == 1:
            continue  # header
        if not row:
            continue
        q = row[0] if len(row) > 0 else None
        if not q or not str(q).strip():
            continue
        if not ID_PATTERN.search(str(q)):
            continue
        a = row[1] if len(row) > 1 else None
        links = row[2] if len(row) > 2 else None
        yield idx, (q, a, links)


def make_yaml(item: FAQItem) -> str:
    def esc(v: str) -> str:
        v = str(v or "").replace("\\", "\\\\").replace('"', '\\"')
        return v

    lines = ["---"]
    lines.append(f"id: {item.id}")
    lines.append(f"dept: {item.dept}")
    lines.append(f'question: "{esc(item.question)}"')
    lines.append(f"source_file: {item.source_file}")
    lines.append(f"source_sheet: {item.sheet}")
    lines.append(f"source_row: {item.row}")
    if item.links:
        lines.append("links:")
        for L in item.links:
            lines.append(f'  - name: "{esc(L["name"])}"')
            lines.append(f'    url: "{esc(L["url"])}"')
    else:
        lines.append("links: []")
    lines.append("---")
    return "\n".join(lines)


XREF_PATTERN = re.compile(r"\[([A-Z]{2,4}\d{2,4})\]")


def detect_xrefs(text: str, self_id: str, known_ids: set[str]) -> list[str]:
    """掃文字中的 [XXX001] 引用，回傳去重後的有效 ID 清單（排除自己、排除不存在）。"""
    if not text:
        return []
    found = []
    seen = set()
    for m in XREF_PATTERN.finditer(text):
        rid = m.group(1)
        if rid == self_id or rid in seen or rid not in known_ids:
            continue
        seen.add(rid)
        found.append(rid)
    return found


def render_markdown(item: FAQItem, xrefs: list[str]) -> str:
    body_parts = [make_yaml(item), ""]
    body_parts.append(f"# {item.question}")
    body_parts.append("")
    body_parts.append("## Answer")
    body_parts.append("")
    body_parts.append(item.answer.strip() if item.answer else "_（無內容）_")
    if item.links:
        body_parts.append("")
        body_parts.append("## Links")
        body_parts.append("")
        for L in item.links:
            if L["url"]:
                body_parts.append(f"- [{L['name']}]({L['url']})")
            else:
                body_parts.append(f"- {L['name']}")
    if xrefs:
        body_parts.append("")
        body_parts.append("## Related")
        body_parts.append("")
        for rid in xrefs:
            body_parts.append(f"- [[{rid}]]")
    body_parts.append("")
    return "\n".join(body_parts)


def convert_one_file(xlsx_path: Path) -> list[FAQItem]:
    filename = xlsx_path.name
    dept = detect_dept(filename)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    sheet = pick_sheet(wb, filename)
    ws = wb[sheet]

    items: list[FAQItem] = []
    for row_idx, (q, a, links) in iter_faq_rows(ws):
        qid, title = extract_id_and_title(q)
        if not qid:
            continue
        items.append(
            FAQItem(
                dept=dept,
                id=qid,
                question=title or str(q).strip(),
                answer=str(a or "").strip(),
                links=parse_links(links),
                source_file=filename,
                sheet=sheet,
                row=row_idx,
            )
        )
    return items


DEPT_DESC = {
    "CFC": "財務（會計、出納、報帳、預算、發票）",
    "GAC": "行政（庶務、採購、會議室、識別證、文具）",
    "HQ": "總部 / 高階管理（PFL 統一作業、ATS、BUBP）",
    "IPC": "採購（採購契約、供應商管理、選商驗收）",
    "KMC": "知識管理（檔案命名、知識分享、AAR、ISO 文管）",
    "LCC": "法務（契約管理、法律文件審核、簽約用印）",
    "PCC": "公司治理（CIS、對外訊息、貴賓接待）",
    "PMC": "專案管理（任務、提案、Pipeline、CEM）",
    "TMC": "人資（薪酬 CB、員工關係 ER、招募、出勤）",
}


def write_vault(items: list[FAQItem], out_dir: Path) -> None:
    """寫整套 karpathy-wiki 結構：index.md / log.md / <dept>/_index.md / <dept>/*.md"""
    out_dir.mkdir(parents=True, exist_ok=True)
    known_ids = {it.id for it in items}

    # 1) 寫 per-Q .md（含 Related 跨頁）
    by_dept: dict[str, list[FAQItem]] = {}
    written = 0
    for it in items:
        dept_dir = out_dir / it.dept
        dept_dir.mkdir(exist_ok=True)
        xrefs = detect_xrefs(f"{it.question}\n{it.answer}", it.id, known_ids)
        slug = slugify(it.question) or "untitled"
        fname = f"[{it.id}]-{slug}.md"
        (dept_dir / fname).write_text(render_markdown(it, xrefs), encoding="utf-8")
        by_dept.setdefault(it.dept, []).append(it)
        written += 1
    print(f"  → wrote {written} FAQ files")

    # 2) 寫每部門 _index.md
    for dept, dept_items in sorted(by_dept.items()):
        dept_items.sort(key=lambda x: x.id)
        lines = [
            f"# {dept} 部門 FAQ 目錄",
            "",
            f"_{DEPT_DESC.get(dept, '')}_",
            "",
            f"共 **{len(dept_items)}** 則 FAQ。",
            "",
            "| ID | 問題 |",
            "|---|---|",
        ]
        for it in dept_items:
            short_q = it.question.split("\n")[0][:80]
            lines.append(f"| [[{it.id}]] | {short_q} |")
        lines.append("")
        (out_dir / dept / "_index.md").write_text("\n".join(lines), encoding="utf-8")
    print(f"  → wrote {len(by_dept)} dept _index.md")

    # 3) 寫根 index.md
    total = len(items)
    root_lines = [
        "# iASK 2.0 FAQ Wiki",
        "",
        f"_集團內部知識庫 — 共 **{total}** 則 FAQ，跨 **{len(by_dept)}** 個部門。_",
        "",
        f"Last update: `{_dt.date.today().isoformat()}`",
        "",
        "## 部門",
        "",
        "| 部門 | 範疇 | Q 數 |",
        "|---|---|---|",
    ]
    for dept in sorted(by_dept.keys()):
        n = len(by_dept[dept])
        desc = DEPT_DESC.get(dept, "")
        root_lines.append(f"| [{dept}](./{dept}/_index.md) | {desc} | {n} |")

    # concepts 區塊
    concepts_dir = out_dir / "concepts"
    if concepts_dir.exists():
        concept_files = sorted(concepts_dir.glob("*.md"))
        if concept_files:
            root_lines.extend(["", "## 跨部門概念頁（concepts/）", "",
                "_合成自多部門 FAQ，適合「不知道該找哪個部門」的問題從這裡入口。_", ""])
            for cf in concept_files:
                title = cf.stem
                # 嘗試從檔內 # 標題抽
                try:
                    first_h1 = next((ln[2:].strip() for ln in cf.read_text(encoding="utf-8").splitlines() if ln.startswith("# ")), title)
                    title = first_h1
                except Exception:
                    pass
                root_lines.append(f"- [{title}](./concepts/{cf.name})")

    root_lines.extend([
        "",
        "## 使用方式",
        "",
        "- 直接在 Obsidian 開啟此 vault，或交給 LLM 作為知識庫 context。",
        "- 每個 FAQ 檔附 `[[wikilink]]` 跨頁引用，可在 Obsidian 看 graph。",
        "- 來源：`各單位FAQ/*.xlsx`（部門維護），透過 `scripts/convert_excel_to_vault.py` 自動轉換。",
        "- `concepts/` 是跨部門概念頁，由 LLM 主動合成，可手動微調。",
        "",
    ])
    (out_dir / "index.md").write_text("\n".join(root_lines), encoding="utf-8")
    print("  → wrote index.md")

    # 4) append log.md
    log_path = out_dir / "log.md"
    ts = _dt.datetime.now().strftime("%Y-%m-%d %H:%M")
    log_entry = (
        f"## [{ts}] regenerate | total={total}, depts={len(by_dept)}\n"
        + "".join(f"- {d}: {len(by_dept[d])}\n" for d in sorted(by_dept.keys()))
        + "\n"
    )
    if log_path.exists():
        existing = log_path.read_text(encoding="utf-8")
        log_path.write_text(existing + log_entry, encoding="utf-8")
    else:
        header = "# Wiki Change Log\n\n_Append-only chronological record of vault regenerations._\n\n"
        log_path.write_text(header + log_entry, encoding="utf-8")
    print("  → appended log.md")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", default=str(SOURCE_DIR), help="FAQ Excel folder")
    parser.add_argument("--out", default=str(VAULT_DIR), help="vault output dir")
    args = parser.parse_args()

    src = Path(args.source)
    out = Path(args.out)

    if out.exists():
        # 清空舊內容（保留：.git、log.md append-only、concepts/ 人工/AI 維護）
        preserve = {".git", "log.md", "concepts"}
        for sub in out.iterdir():
            if sub.name.startswith(".") or sub.name in preserve:
                continue
            if sub.is_dir():
                import shutil

                shutil.rmtree(sub)
            else:
                sub.unlink()

    all_items: list[FAQItem] = []
    xlsx_files = sorted(src.glob("FAQ更新作業*iASK*.xlsx"))
    if not xlsx_files:
        print(f"no Excel files found in {src}", file=sys.stderr)
        return 1

    print(f"Converting {len(xlsx_files)} Excel files...")
    for f in xlsx_files:
        items = convert_one_file(f)
        print(f"  {f.name}  →  {len(items)} items")
        all_items.extend(items)

    write_vault(all_items, out)

    # 統計
    print("\n=== Summary ===")
    by_dept: dict[str, int] = {}
    for it in all_items:
        by_dept[it.dept] = by_dept.get(it.dept, 0) + 1
    for d, n in sorted(by_dept.items()):
        print(f"  {d}: {n}")
    print(f"  TOTAL: {len(all_items)}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
